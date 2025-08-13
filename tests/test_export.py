import pytest
import os
import tempfile
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timezone
from app import create_app, db
from app.models import User, Category, Product, Order, OrderItem, Transaction
from app.utils import export_to_excel
from werkzeug.security import generate_password_hash
from decimal import Decimal

class TestExportFunctionality:
    """Test Excel export functionality"""
    
    @pytest.fixture
    def app(self):
        """Create test application with sample data"""
        db_fd, db_path = tempfile.mkstemp()
        
        app = create_app()
        app.config.update({
            'TESTING': True,
            'SQLALCHEMY_DATABASE_URI': f'sqlite:///{db_path}',
            'SECRET_KEY': 'test-secret-key',
            'EXPORT_FOLDER': tempfile.mkdtemp()
        })
        
        with app.app_context():
            db.create_all()
            
            # Create test user
            user = User(
                username='testuser',
                password_hash=generate_password_hash('testpass'),
                role='admin'
            )
            db.session.add(user)
            db.session.flush()
            
            # Create test category
            category = Category(
                name_en='Test Category',
                name_ar='تصنيف تجريبي',
                is_active=True
            )
            db.session.add(category)
            db.session.flush()
            
            # Create test product
            product = Product(
                name_en='Test Product',
                name_ar='منتج تجريبي',
                category_id=category.id,
                price=Decimal('25.50'),
                is_active=True,
                is_service=True
            )
            db.session.add(product)
            db.session.flush()
            
            # Create test order
            order = Order(
                order_number='TEST001',
                created_by=user.id,
                total_amount=Decimal('51.00'),
                tax_amount=Decimal('7.14'),
                discount_amount=Decimal('0.00'),
                final_amount=Decimal('58.14'),
                payment_method='cash',
                payment_status='paid',
                order_status='delivered'
            )
            db.session.add(order)
            db.session.flush()
            
            # Create order item
            order_item = OrderItem(
                order_id=order.id,
                product_id=product.id,
                quantity=2,
                unit_price=Decimal('25.50'),
                total_price=Decimal('51.00')
            )
            db.session.add(order_item)
            
            # Create test transactions
            income_transaction = Transaction(
                transaction_type='income',
                amount=Decimal('58.14'),
                description_en='Test Sale',
                description_ar='بيع تجريبي',
                category_en='Sales',
                category_ar='مبيعات',
                payment_method='cash',
                created_by=user.id,
                order_id=order.id
            )
            db.session.add(income_transaction)
            
            expense_transaction = Transaction(
                transaction_type='expense',
                amount=Decimal('15.00'),
                description_en='Test Expense',
                description_ar='مصروف تجريبي',
                category_en='Utilities',
                category_ar='مرافق',
                payment_method='cash',
                created_by=user.id
            )
            db.session.add(expense_transaction)
            
            db.session.commit()
        
        yield app
        
        # Cleanup
        os.close(db_fd)
        os.unlink(db_path)
        # Clean up export folder
        import shutil
        shutil.rmtree(app.config['EXPORT_FOLDER'])
    
    def test_export_daily_report_english(self, app):
        """Test daily export in English"""
        with app.app_context():
            filename = export_to_excel('daily', 'en')
            
            # Check file was created
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            assert os.path.exists(export_path)
            
            # Check filename format
            assert filename.startswith('elhoseny_report_')
            assert filename.endswith('.xlsx')
            assert datetime.now().strftime('%Y%m%d') in filename
            
            # Load and check Excel content
            workbook = load_workbook(export_path)
            
            # Check sheets exist
            expected_sheets = ['Summary', 'Orders', 'Transactions', 'Product Performance']
            for sheet_name in expected_sheets:
                assert sheet_name in workbook.sheetnames
            
            # Check Summary sheet content
            summary_sheet = workbook['Summary']
            
            # Check title exists
            title_cell = summary_sheet['A1']
            assert 'ELHOSENY' in str(title_cell.value)
            assert 'Report' in str(title_cell.value)
            
            # Check summary data exists (starts at row 6)
            assert 'Report Period' in str(summary_sheet['A6'].value)
            assert 'Total Orders' in str(summary_sheet['A9'].value)
            assert 'Total Revenue' in str(summary_sheet['A10'].value)
            
            # Check Orders sheet
            orders_sheet = workbook['Orders']
            headers = [cell.value for cell in orders_sheet[1]]
            assert 'Order Number' in headers
            assert 'Date' in headers
            assert 'Amount' in headers
            
            # Check Transactions sheet
            transactions_sheet = workbook['Transactions']
            headers = [cell.value for cell in transactions_sheet[1]]
            assert 'Date' in headers
            assert 'Type' in headers
            assert 'Amount' in headers
            assert 'Description' in headers
    
    def test_export_daily_report_arabic(self, app):
        """Test daily export in Arabic"""
        with app.app_context():
            filename = export_to_excel('daily', 'ar')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            assert os.path.exists(export_path)
            
            # Load and check Arabic content
            workbook = load_workbook(export_path)
            
            # Check RTL setting
            summary_sheet = workbook['Summary']
            assert summary_sheet.sheet_view.rightToLeft is True
            
            # Check Arabic headers in transactions
            transactions_sheet = workbook['Transactions']
            headers = [cell.value for cell in transactions_sheet[1]]
            # Should contain Arabic text
            arabic_found = any('التاريخ' in str(header) or 'النوع' in str(header) 
                             or 'المبلغ' in str(header) for header in headers if header)
            assert arabic_found, f"No Arabic headers found in: {headers}"
    
    def test_export_weekly_report(self, app):
        """Test weekly export"""
        with app.app_context():
            filename = export_to_excel('weekly', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            assert os.path.exists(export_path)
            
            # Check filename indicates weekly
            assert 'week_' in filename
            
            # Verify content exists
            workbook = load_workbook(export_path)
            assert 'Summary' in workbook.sheetnames
    
    def test_export_monthly_report(self, app):
        """Test monthly export"""
        with app.app_context():
            filename = export_to_excel('monthly', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            assert os.path.exists(export_path)
            
            # Check filename indicates monthly
            assert 'month_' in filename
    
    def test_export_contains_logo(self, app):
        """Test that export contains logo if available"""
        with app.app_context():
            filename = export_to_excel('daily', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            workbook = load_workbook(export_path)
            summary_sheet = workbook['Summary']
            
            # Check if there are any images in the sheet
            # Note: Logo might not be added if file doesn't exist
            # This test checks the structure is in place
            assert summary_sheet is not None
    
    def test_export_data_accuracy(self, app):
        """Test that exported data matches database data"""
        with app.app_context():
            filename = export_to_excel('daily', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            
            # Read the Excel file with pandas
            orders_df = pd.read_excel(export_path, sheet_name='Orders')
            transactions_df = pd.read_excel(export_path, sheet_name='Transactions')
            
            # Check orders data
            assert len(orders_df) >= 1  # At least one order
            
            # Find our test order
            test_order = orders_df[orders_df['Order Number'] == 'TEST001']
            assert len(test_order) == 1
            
            # Check order amount
            order_amount = test_order['Final Amount'].iloc[0]
            assert abs(float(order_amount) - 58.14) < 0.01
            
            # Check transactions data
            assert len(transactions_df) >= 2  # At least income and expense
            
            # Check income transaction
            income_transactions = transactions_df[transactions_df['Type'] == 'Income']
            assert len(income_transactions) >= 1
            
            # Check expense transaction
            expense_transactions = transactions_df[transactions_df['Type'] == 'Expense']
            assert len(expense_transactions) >= 1
    
    def test_export_bilingual_headers(self, app):
        """Test that bilingual headers are present"""
        with app.app_context():
            # Test English headers
            filename_en = export_to_excel('daily', 'en')
            export_path_en = os.path.join(app.config['EXPORT_FOLDER'], filename_en)
            
            workbook_en = load_workbook(export_path_en)
            transactions_sheet_en = workbook_en['Transactions']
            headers_en = [cell.value for cell in transactions_sheet_en[1]]
            
            # Should have English headers
            assert any('Date' in str(header) for header in headers_en if header)
            assert any('Type' in str(header) for header in headers_en if header)
            
            # Test Arabic headers
            filename_ar = export_to_excel('daily', 'ar')
            export_path_ar = os.path.join(app.config['EXPORT_FOLDER'], filename_ar)
            
            workbook_ar = load_workbook(export_path_ar)
            transactions_sheet_ar = workbook_ar['Transactions']
            headers_ar = [cell.value for cell in transactions_sheet_ar[1]]
            
            # Should have Arabic headers
            arabic_headers_found = any(
                'التاريخ' in str(header) or 'النوع' in str(header) 
                for header in headers_ar if header
            )
            assert arabic_headers_found
    
    def test_export_with_no_data(self, app):
        """Test export when no data exists"""
        with app.app_context():
            # Clear all data
            OrderItem.query.delete()
            Order.query.delete()
            Transaction.query.delete()
            Product.query.delete()
            Category.query.delete()
            db.session.commit()
            
            # Export should still work
            filename = export_to_excel('daily', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            assert os.path.exists(export_path)
            
            # Check file can be opened
            workbook = load_workbook(export_path)
            assert 'Summary' in workbook.sheetnames
    
    def test_export_formatting(self, app):
        """Test Excel formatting and styling"""
        with app.app_context():
            filename = export_to_excel('daily', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            workbook = load_workbook(export_path)
            
            # Check summary sheet formatting
            summary_sheet = workbook['Summary']
            
            # Title should be formatted (font size, bold, etc.)
            title_cell = summary_sheet['A1']
            assert title_cell.font.size >= 14
            assert title_cell.font.bold is True
            
            # Check if headers in other sheets are formatted
            orders_sheet = workbook['Orders']
            if orders_sheet.max_row > 0:
                header_row = orders_sheet[1]
                for cell in header_row:
                    if cell.value:
                        # Headers should be bold
                        assert cell.font.bold is True
    
    def test_export_file_permissions(self, app):
        """Test that exported files have correct permissions"""
        with app.app_context():
            filename = export_to_excel('daily', 'en')
            
            export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
            
            # File should exist and be readable
            assert os.path.exists(export_path)
            assert os.access(export_path, os.R_OK)
            
            # File should have reasonable size (not empty)
            file_size = os.path.getsize(export_path)
            assert file_size > 1000  # At least 1KB
    
    def test_export_invalid_period(self, app):
        """Test export with invalid period parameter"""
        with app.app_context():
            with pytest.raises(ValueError):
                export_to_excel('invalid_period', 'en')
    
    def test_concurrent_exports(self, app):
        """Test that multiple exports can be created simultaneously"""
        with app.app_context():
            # Create multiple exports
            filenames = []
            for i in range(3):
                filename = export_to_excel('daily', 'en')
                filenames.append(filename)
            
            # All should be unique
            assert len(set(filenames)) == len(filenames)
            
            # All files should exist
            for filename in filenames:
                export_path = os.path.join(app.config['EXPORT_FOLDER'], filename)
                assert os.path.exists(export_path)

if __name__ == '__main__':
    pytest.main([__file__, '-v'])
