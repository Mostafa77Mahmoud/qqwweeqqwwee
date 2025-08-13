import os
import shutil
import logging
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import func, and_

from app import db
from app.models import Order, Transaction, BackupLog, Settings

logger = logging.getLogger(__name__)

def generate_order_number():
    """Generate unique order number"""
    today = datetime.utcnow()
    date_str = today.strftime('%Y%m%d')
    
    # Count orders for today
    today_count = Order.query.filter(
        func.date(Order.created_at) == today.date()
    ).count() + 1
    
    return f"ORD-{date_str}-{today_count:04d}"

def export_to_excel(period='daily', language='en'):
    """Export data to Excel with bilingual headers and logo"""
    try:
        # Ensure exports directory exists
        os.makedirs('exports', exist_ok=True)
        
        # Calculate date range based on period
        end_date = datetime.utcnow()
        
        if period == 'daily':
            start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'weekly':
            start_date = end_date - timedelta(days=7)
        elif period == 'monthly':
            start_date = end_date - timedelta(days=30)
        else:
            start_date = end_date - timedelta(days=1)
        
        # Get data
        transactions = Transaction.query.filter(
            Transaction.created_at >= start_date
        ).order_by(Transaction.created_at.desc()).all()
        
        orders = Order.query.filter(
            Order.created_at >= start_date
        ).order_by(Order.created_at.desc()).all()
        
        # Prepare bilingual headers
        if language == 'ar':
            headers = {
                'date': 'التاريخ',
                'type': 'النوع',
                'amount': 'المبلغ',
                'description': 'الوصف',
                'payment_method': 'طريقة الدفع',
                'order_number': 'رقم الطلب',
                'customer': 'العميل',
                'status': 'الحالة',
                'total': 'الإجمالي'
            }
        else:
            headers = {
                'date': 'Date',
                'type': 'Type',
                'amount': 'Amount',
                'description': 'Description',
                'payment_method': 'Payment Method',
                'order_number': 'Order Number',
                'customer': 'Customer',
                'status': 'Status',
                'total': 'Total'
            }
        
        # Create filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'exports/elhoseny_report_{period}_{timestamp}.xlsx'
        
        # Create Excel workbook with multiple sheets
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Overview sheet
            overview_data = []
            
            # Calculate totals
            total_income = sum(t.amount for t in transactions if t.type == 'income')
            total_expense = sum(t.amount for t in transactions if t.type == 'expense')
            net_income = total_income - total_expense
            
            if language == 'ar':
                overview_data = [
                    ['الفترة', f'{start_date.strftime("%Y-%m-%d")} إلى {end_date.strftime("%Y-%m-%d")}'],
                    ['إجمالي الدخل', f'{total_income:.2f}'],
                    ['إجمالي المصروفات', f'{total_expense:.2f}'],
                    ['صافي الدخل', f'{net_income:.2f}'],
                    ['عدد الطلبات', len(orders)],
                    ['الطلبات المكتملة', len([o for o in orders if o.status == 'completed'])],
                    ['الطلبات المعلقة', len([o for o in orders if o.status == 'pending'])]
                ]
            else:
                overview_data = [
                    ['Period', f'{start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'],
                    ['Total Income', f'{total_income:.2f}'],
                    ['Total Expense', f'{total_expense:.2f}'],
                    ['Net Income', f'{net_income:.2f}'],
                    ['Total Orders', len(orders)],
                    ['Completed Orders', len([o for o in orders if o.status == 'completed'])],
                    ['Pending Orders', len([o for o in orders if o.status == 'pending'])]
                ]
            
            overview_df = pd.DataFrame(overview_data, columns=['Metric', 'Value'])
            overview_df.to_excel(writer, sheet_name='Overview', index=False)
            
            # Transactions sheet
            transaction_data = []
            for t in transactions:
                description = t.get_description(language)
                transaction_data.append([
                    t.created_at.strftime('%Y-%m-%d %H:%M'),
                    t.type,
                    float(t.amount),
                    description,
                    t.payment_method or '',
                    t.category or ''
                ])
            
            if transaction_data:
                transaction_df = pd.DataFrame(transaction_data, columns=[
                    headers['date'],
                    headers['type'],
                    headers['amount'],
                    headers['description'],
                    headers['payment_method'],
                    'Category'
                ])
                transaction_df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Orders sheet
            order_data = []
            for o in orders:
                customer_name = o.customer.name if o.customer else ''
                order_data.append([
                    o.created_at.strftime('%Y-%m-%d %H:%M'),
                    o.order_number,
                    customer_name,
                    float(o.total_amount),
                    o.payment_method,
                    o.status
                ])
            
            if order_data:
                order_df = pd.DataFrame(order_data, columns=[
                    headers['date'],
                    headers['order_number'],
                    headers['customer'],
                    headers['total'],
                    headers['payment_method'],
                    headers['status']
                ])
                order_df.to_excel(writer, sheet_name='Orders', index=False)
        
        # Add logo and formatting
        try:
            workbook = load_workbook(filename)
            
            # Add logo to overview sheet if it exists
            if 'Overview' in workbook.sheetnames:
                overview_sheet = workbook['Overview']
                
                logo_path = 'app/static/images/elhoseny_logo.jpg'
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    img.width = 100
                    img.height = 100
                    overview_sheet.add_image(img, 'D1')
                
                # Format headers
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='2E5BBA', end_color='2E5BBA', fill_type='solid')
                
                for cell in overview_sheet[1]:
                    if cell.value:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
            
            workbook.save(filename)
            
        except Exception as e:
            logger.warning(f"Could not add logo or formatting to Excel: {e}")
        
        logger.info(f"Excel export created: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise

def create_backup():
    """Create database backup"""
    try:
        # Ensure backups directory exists
        os.makedirs('backups', exist_ok=True)
        
        # Create backup filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        
        # For SQLite, copy the database file
        if 'sqlite' in db.engine.url.drivername:
            db_path = db.engine.url.database
            backup_filename = f'backups/laundry_pos_backup_{timestamp}.db'
            shutil.copy2(db_path, backup_filename)
        else:
            # For other databases, you would use database-specific backup commands
            # This is a placeholder for PostgreSQL or other databases
            backup_filename = f'backups/laundry_pos_backup_{timestamp}.sql'
            # TODO: Implement PostgreSQL backup using pg_dump
            raise NotImplementedError("Backup for non-SQLite databases not implemented")
        
        # Get file size
        file_size = os.path.getsize(backup_filename)
        
        # Log backup
        backup_log = BackupLog(
            filename=os.path.basename(backup_filename),
            file_size=file_size,
            status='success'
        )
        db.session.add(backup_log)
        db.session.commit()
        
        # Clean old backups (keep last 30 days)
        cleanup_old_backups()
        
        logger.info(f"Database backup created: {backup_filename}")
        return backup_filename
        
    except Exception as e:
        # Log failed backup
        backup_log = BackupLog(
            filename=f'backup_{timestamp}',
            status='failed',
            error_message=str(e)
        )
        db.session.add(backup_log)
        db.session.commit()
        
        logger.error(f"Error creating backup: {e}")
        raise

def cleanup_old_backups(retention_days=30):
    """Clean up old backup files"""
    try:
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        
        # Get old backup logs
        old_backups = BackupLog.query.filter(
            BackupLog.created_at < cutoff_date
        ).all()
        
        for backup in old_backups:
            # Delete file if it exists
            backup_path = f'backups/{backup.filename}'
            if os.path.exists(backup_path):
                os.remove(backup_path)
                logger.info(f"Deleted old backup: {backup_path}")
            
            # Delete log entry
            db.session.delete(backup)
        
        db.session.commit()
        
    except Exception as e:
        logger.error(f"Error cleaning up old backups: {e}")

def get_dashboard_stats(days=7):
    """Get dashboard statistics for the last N days"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    # Daily income/expense
    daily_stats = []
    for i in range(days):
        date = start_date + timedelta(days=i)
        date_only = date.date()
        
        day_income = db.session.query(func.sum(Transaction.amount)).filter(
            and_(
                Transaction.type == 'income',
                func.date(Transaction.created_at) == date_only
            )
        ).scalar() or 0
        
        day_expense = db.session.query(func.sum(Transaction.amount)).filter(
            and_(
                Transaction.type == 'expense',
                func.date(Transaction.created_at) == date_only
            )
        ).scalar() or 0
        
        daily_stats.append({
            'date': date_only.isoformat(),
            'income': float(day_income),
            'expense': float(day_expense),
            'net': float(day_income - day_expense)
        })
    
    return daily_stats

def validate_file_upload(file, allowed_extensions=None, max_size=None):
    """Validate uploaded file"""
    if allowed_extensions is None:
        allowed_extensions = {'jpg', 'jpeg', 'png', 'gif'}
    
    if max_size is None:
        max_size = 5 * 1024 * 1024  # 5MB
    
    if not file or not file.filename:
        return False, "No file selected"
    
    # Check file extension
    if '.' not in file.filename:
        return False, "No file extension"
    
    ext = file.filename.rsplit('.', 1)[1].lower()
    if ext not in allowed_extensions:
        return False, f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
    
    # Check file size (if we can get it)
    try:
        file.seek(0, 2)  # Seek to end
        size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if size > max_size:
            return False, f"File too large. Maximum size: {max_size // (1024*1024)}MB"
    except:
        pass  # Can't check size, continue
    
    return True, "File is valid"

def format_currency(amount, currency_code='EGP'):
    """Format currency amount"""
    try:
        settings = Settings.query.first()
        if settings:
            symbol = settings.currency_symbol
        else:
            symbol = 'ج.م' if currency_code == 'EGP' else currency_code
        
        return f"{symbol} {amount:,.2f}"
    except:
        return f"{amount:,.2f}"

def get_system_info():
    """Get system information for debugging"""
    import platform
    import sys
    
    return {
        'platform': platform.platform(),
        'python_version': sys.version,
        'app_version': '1.0.0',
        'database_url': db.engine.url.__to_string__(hide_password=True),
        'backup_count': BackupLog.query.count(),
        'user_count': db.session.query(func.count(db.distinct(db.text('users.id')))).scalar(),
        'order_count': Order.query.count(),
        'transaction_count': Transaction.query.count()
    }
